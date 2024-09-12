"""
Excel VBA - Double Click
https://www.youtube.com/watch?v=vFd2vZw4yJ0

Execute Shell Commands from Excel Cell
https://superuser.com/questions/1220696/execute-shell-commands-from-excel-cell

Python : Reading Large Excel Worksheets using Openpyxl
https://stackoverflow.com/questions/31189727/python-reading-large-excel-worksheets-using-openpyxl
"""

import argparse
import os
import sys
import webbrowser
from itertools import groupby
from pathlib import Path

import openpyxl
from jinja2 import Template
from loguru import logger


ROOT_DIR = Path(os.path.dirname(sys.argv[0]))
LOGGER_PATH = ROOT_DIR / "conceptz.log"
TEMPLATE_PATH = ROOT_DIR / "template.html"
HTML_DIR = ROOT_DIR / "conceptz_html"

CONCEPTS_SHEET_IDX = 0
INFO_SHEET_IDX = 1


logger.add(LOGGER_PATH, format="{time} | {level} | {message}", level="INFO")
os.makedirs(HTML_DIR, exist_ok=True)


@logger.catch
def main():

    parser = argparse.ArgumentParser(description='Conceptz for Excel')
    parser.add_argument("xls_path", type=str, help="Data file")
    parser.add_argument("concept_name", type=str, help="Concept name")
    parser.add_argument("--read_only", type=bool, default=False, help="Open xls in read only mode")
    args = parser.parse_args()
    logger.info(args.xls_path)
    logger.info(args.concept_name)

    xls_path = Path(args.xls_path)
    concept_name = args.concept_name
    read_only = args.read_only

    wb = openpyxl.load_workbook(str(xls_path), read_only=read_only)

    # сперва надо найти строку с концептом
    concept_row = None

    # искать его будем в листах, начинающихся на "Concepts"
    for ws in wb.worksheets:

        if not ws.title.startswith("Concepts"):
            continue
        if ws.max_row == 0:
            continue

        # определяем названия столбцов
        concept_keys = [ws.cell(1, column).value for column in range(1, ws.max_column + 1)]

        # определяем в каком столбце хранится concept.id
        concept_id_column = concept_keys.index("ID") + 1
        if concept_id_column is None:
            raise Exception(f'Column "ID" not found in worksheet "{ws.title}"')

        # определяем в каком столбце хранится concept.name
        concept_name_column = concept_keys.index("Name") + 1
        if concept_name_column is None:
            raise Exception(f'Column "Name" not found in worksheet "{ws.title}"')

        # ищем строку, совпадающую по имени
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, concept_name_column).value == concept_name:
                concept_row = {
                    key: ws.cell(row, column).value
                    for column, key in enumerate(concept_keys, start=1)
                    if key is not None
                }
                break

        if concept_row is not None:
            break

    if concept_row is None:
        raise Exception("Concept row not found")

    # теперь, зная id концепта, можно найти информацию по нему
    concept_id = str(concept_row["ID"])
    info_rows = []

    # искать инфу будем в листах, начинающихся на "Info"
    for ws in wb.worksheets:

        if not ws.title.startswith("Info"):
            continue
        if ws.max_row == 0:
            continue

        # определяем названия столбцов
        info_keys = [ws.cell(1, column).value for column in range(1, ws.max_column + 1)]

        # определяем в каком столбце хранится info.concept_id
        concept_id_column = info_keys.index("Concept ID") + 1
        if concept_id_column is None:
            raise Exception(f'Column "Concept ID" not found in worksheet "{ws.title}"')

        # ищем строки, совпадающие по Concept ID
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row, concept_id_column).value) == concept_id:
                info_row = {
                    key: ws.cell(row, column).value
                    for column, key in enumerate(info_keys, start=1)
                    if key is not None
                }
                info_rows.append(info_row)

    # делаем прямые ссылки на скриншоты
    for row in info_rows:
        screenshot = row.get("Screenshot")
        if not screenshot:
            continue
        if "mail.ru" not in screenshot:
            continue
        try:
            xxxx, yyyyyyyyy = screenshot.split('/')[-2:]
        except Exception:
            logger.info(f"Can't parse screenshot link: {screenshot}")
        else:
            row["Screenshot"] = f"https://thumb.cloud.mail.ru/weblink/thumb/xw1/{xxxx}/{yyyyyyyyy}"

    # учитываем таймкоды
    for row in info_rows:
        if row["Timecode"]:
            row["Source link"] = row["Source link"] + "?t=" + row["Timecode"]
        del row["Timecode"]

    # группируем информацию
    sources = [ir for ir in info_rows if ir["Type"] == "source"]
    examples = [ir for ir in info_rows if ir["Type"] == "example"]
    screenshots = [ir for ir in info_rows if ir["Type"] == "screenshot"]
    notes = [ir for ir in info_rows if ir["Type"] == "note"]
    notes_groups = [
        {
            "Source name": key,
            "Notes": list(group)
        }
        for key, group in groupby(notes, key=lambda x: x["Source name"])
    ]

    # делаем html с помощью jinja2
    with open(TEMPLATE_PATH, "r", encoding="utf-8") as file:
        template_html = file.read()

    template = Template(template_html)
    rendered_html = template.render(
        concept=concept_row,
        sources=sources,
        examples=examples,
        screenshots=screenshots,
        notes=notes_groups
    )
    html_path = HTML_DIR / f"{concept_name}.html"

    with open(html_path, "w", encoding="utf-8") as file:
        file.write(rendered_html)

    # открываем полученный из шаблона html
    webbrowser.open(str(html_path), new=2)


if __name__ == '__main__':
    main()


# debug
# C:\Users\Denis\PythonProjects\conceptz_excel\venv\Scripts\python C:\Users\Denis\PythonProjects\conceptz_excel\conceptz.py C:\Users\Denis\Desktop\CONCEPTZ\conceptz.xlsm Orchestration
